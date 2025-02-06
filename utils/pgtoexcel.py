import openpyxl
from openpyxl.styles import Font


async def export_to_excel(data, headings, filepath):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.row_dimensions[1].font = Font(bold=True)

    for colno, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=colno).value = heading

    # Debugging: Check data being written to the Excel file
    print(f"Writing the following data to Excel: {data}")

    for rowno, row in enumerate(data, start=2):
        for colno, cell_value in enumerate(row, start=1):
            sheet.cell(row=rowno, column=colno).value = cell_value

    wb.save(filepath)
    print(f"Excel file saved at {filepath}")
